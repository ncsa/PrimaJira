import openpyxl
import configparser
import main as m
import shlog


def user_id_from_name(serv, usr, passw, name):
    if name:  # if it's not None
        name = name.replace(',', '')
        name_first = name.split(' ')[0]
        name_last = name.split(' ')[-1]
        request_data = {'Field': ['ObjectId', 'Name']
            , 'Filter': "PersonalName like '%%%s%%' and PersonalName like '%%%s%%'" % (name_first, name_last)
                        }
        synched = m.soap_request(request_data, serv, 'UserService', 'ReadUsers', usr, passw)
        return synched[0]['ObjectId'], synched[0]['Name']
    else:
        return None, None


def post_note(serv, usr, passw, parent, code, note):
    request_data = {'ActivityNote': {'Note': note,
                                     'ActivityObjectId': parent,
                                     'NotebookTopicObjectId': code  # 43 and 38 this better never change
                                     }}
    synched = m.soap_request(request_data, serv, 'ActivityNoteService', 'CreateActivityNotes', usr, passw)
    return synched


def resource_id_from_name(serv, usr, passw, name):
    name = name.replace(',', '')
    name_first = name.split(' ')[0]
    name_last = name.split(' ')[-1]
    request_data = {'Field': ['ObjectId','Id']
        , 'Filter': "Name like '%%%s%%' and Name like '%%%s%%' and ResourceType = 'Labor' and Id not like 'RS%%'"
                    "" % (name_first, name_last)
                    }
    if name:  # if it's not None
        synched = m.soap_request(request_data, serv, 'ResourceService', 'ReadResources', usr, passw)
        try:
            return synched[0]['ObjectId']
        except IndexError:
            # name not found in resources
            return None
    else:
        return None


def post_resource_assign(serv, usr, passw, activity, name):
    resource = resource_id_from_name(serv, usr, passw, name)
    if resource:
        # check if the resource had already been assigned
        request_data = {'Field': ['ResourceObjectId'],
                        'Filter': "ActivityObjectId = '%s' and ResourceObjectId = '%s'" % (str(activity), str(resource))}
        resp = m.soap_request(request_data, primaserver, 'ResourceAssignmentService', 'ReadResourceAssignments', primauser,
                              primapasswd)
        if len(resp) > 0:
            # already assigned
            synched = 'already assigned to activity #' + str(activity)
            shlog.verbose(name + ' ' + synched)
        else:
            # actual post
            request_data = {'ResourceAssignment': {'ActivityObjectId': activity,
                                                   'ResourceObjectId': resource}}
            synched = m.soap_request(request_data, serv, 'ResourceAssignmentService', 'CreateResourceAssignments', usr, passw)
    else:
        return name + ' not found in resources'
    return synched


def post_step_resource_assign(serv, usr, passw, step, name):
    resource = resource_id_from_name(serv, usr, passw, name)
    if resource:
        # check if the resource had already been assigned
        name = name.replace(',', '')
        name_first = name.split(' ')[0]
        name_last = name.split(' ')[-1]
        request_data = {
            'Field': ['Text'],
            'Filter': "UDFTypeObjectId = '158' and ForeignObjectId = '%s' and Text like '%%%s%%' and "
                      "Text like '%%%s%%'" % (str(step), str(name_first), str(name_last))}
        resp = m.soap_request(request_data, primaserver, 'UDFValueService', 'ReadUDFValues', primauser, primapasswd)
        if len(resp) > 0:
            # already assigned
            synched = 'already assigned to step #' + str(step)
            shlog.verbose(name + ' ' + synched)
        else:
            # actual post
            try:
                request_data = {'Field': ['Name'],
                                'Filter': "ObjectId = '%s'" % str(resource)
                               }
                synched = m.soap_request(request_data, serv, 'ResourceService', 'ReadResources', usr, passw)
                if len(synched) > 0:  # if it's not None
                    p6_name = synched[0]['Name']
                else:
                    p6_name = None

                baseline = m.actual_baseline(serv, usr, passw)
                resp = m.ticket_wipe(primaserver, primauser, primapasswd, step, 158) # just in case
                shlog.verbose('Posting name ' + p6_name + ' to step #' + str(step))
                m.ticket_post(serv, usr, passw, step, baseline, p6_name, 158)
            except IndexError:
                return 'IndexError thrown, user not found'
    else:
        return name + ' not found in resources'
    return synched

def post_color(serv, usr, passw, activity, color, code):
    # add a checkmark to an activity
    # colors available: Yellow, Green
    if color not in ['Yellow','Green']:
        shlog.normal('Color error! ' + str(color) + ' not in list: Yellow, Green')
        exit(0)
    request_data = {'UDFValue': {'ForeignObjectId': activity,
                                 'ProjectObjectId': m.actual_baseline(serv, usr, passw),
                                 'Indicator': color,
                                 'UDFTypeObjectId': code}}
    # Import into NCSA JIRA is 153
    # Import into LSST JIRA is 148
    synched = m.soap_request(request_data, serv, 'UDFValueService', 'CreateUDFValues', usr, passw)
    shlog.verbose('Server response: ' + str(synched))
    return synched

def remove_colors(serv, usr, passw, activity, code):
    # use ticket_wipe to make udf delete calls
    resp = m.ticket_wipe(serv, usr, passw, activity, code)
    return resp


# read config and get primavera login info
parser = configparser.ConfigParser()
with open('login') as configfile:
    parser.read_file(configfile)
primadict=parser['primavera-section']
primauser=primadict['user']
primapasswd=primadict['passwd']
primaserver=primadict['server']
# read tool config
tool_dict = parser['tool-settings']
tool_log = tool_dict['loglevel']
loglevel=shlog.__dict__[tool_log]
assert type(loglevel) == type(1)
shlog.basicConfig(level=shlog.__dict__[tool_log])
tool_fixer = tool_dict['fix']
if tool_fixer.lower() in ['true', '1', 't', 'y', 'yes', 'yeah', 'yup', 'certainly', 'uh-huh']:
    tool_fixer = True
else:
    tool_fixer = False


if __name__ == '__main__':
    # preload file and worksheet
    workbook = openpyxl.load_workbook('input.xlsx', data_only=True)
    acts_sheet = workbook.get_sheet_by_name('Epic Suggestions')
    steps_sheet = workbook.get_sheet_by_name('TaskStory Suggestion')

    m.vpn_toggle(True)
    # go through activity entries in the file
    # because max_row method counts rows even after they've been wiped, it's time to go back to the old VBA ways...
    i = 2
    while True:
        import_check = acts_sheet.cell(row=i, column=1).value
        act_name = acts_sheet.cell(row=i, column=2).value
        if import_check == 'TRUE' or import_check == True:
            continue  # skip TRUE items
        if import_check == '' or import_check == None or import_check == ' ' or act_name == None:
            break  # stop execution
        shlog.verbose('Processing activity called ' + act_name)
        owner = acts_sheet.cell(row=i, column=3).value
        api_owner_obj, api_owner_name = user_id_from_name(primaserver, primauser, primapasswd, owner)
        purpose = acts_sheet.cell(row=i, column=5).value
        scope = acts_sheet.cell(row=i, column=6).value
        # Activity creation code
        request_data = {'Activity': {'Name': act_name,
                                     'AtCompletionDuration': 0,
                                     'CalendarObjectId': 638,  # NCSA Standard with Holidays, ok to hardcode
                                     'ProjectObjectId': m.actual_baseline(primaserver, primauser, primapasswd),
                                     # 408 is the one currently active
                                     'ProjectId': 'LSST MREFC',
                                     'WBSObjectId': 4597,
                                     'WBSPath': 'LSST MREFC.MREFC.LSST Construction.Test WBS',  # the big dump
                                     'OwnerIDArray': api_owner_obj,
                                     'OwnerNamesArray': api_owner_name
                                     }}
        synched = m.soap_request(request_data, primaserver, 'ActivityService', 'CreateActivities', primauser, primapasswd)
        created_activity = synched[0]
        shlog.normal(str(created_activity) + ' ObjectId returned')
        # post descriptions to the new activity
        # purpose
        synched = post_note(primaserver, primauser, primapasswd, created_activity, 38, purpose)
        # scope
        synched = post_note(primaserver, primauser, primapasswd, created_activity, 43, scope)
        # go through all steps and find relevant ones
        r = 2
        while True:
            step_name = steps_sheet.cell(row=r, column=2).value
            if step_name == '' or step_name == None or step_name == ' ':
                break
            if steps_sheet.cell(row=r, column=1).value.lower() == act_name.lower():
                shlog.verbose('Processing step called ' + step_name)
                step_desc = steps_sheet.cell(row=r, column=7).value
                # assign resources to the parent
                resource = steps_sheet.cell(row=r, column=3).value
                post_resource_assign(primaserver, primauser, primapasswd, created_activity, resource)
                try:
                    step_pts = int(steps_sheet.cell(row=r, column=5).value)/4
                except TypeError:
                    step_pts = 0
                # Step creation code
                request_data = {'ActivityStep': {'ActivityObjectId': created_activity,
                                             'Description': step_desc,
                                             'Name': step_name,
                                             'Weight': step_pts}}
                synched = m.soap_request(request_data, primaserver, 'ActivityStepService', 'CreateActivitySteps', primauser, primapasswd)
                created_step = synched[0]
                post_step_resource_assign(primaserver, primauser, primapasswd, created_step, resource)

            r += 1
        i += 1
        # yellow markings to celebrate completion
        synched = post_color(primaserver, primauser, primapasswd, created_activity, 'Yellow', 153)
        synched = post_color(primaserver, primauser, primapasswd, created_activity, 'Yellow', 148)

    # 41186
    m.vpn_toggle(False)